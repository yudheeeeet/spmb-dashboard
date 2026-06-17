import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_for_school(school_data, output_file):
    info = school_data.get('info', {})
    school_name = info.get('nama_satuan_pendidikan', 'SMP NEGERI')
    npsn = info.get('npsn', 'UNKNOWN')
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=14, bold=True, color="1F497D")
    data_font = Font(name=font_family, size=10)
    bold_data_font = Font(name=font_family, size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    columns = [
        ("No", 6, align_center),
        ("No. Pendaftaran", 20, align_center),
        ("Nama Murid", 25, align_left),
        ("Asal Sekolah", 30, align_left),
        ("Sub Jalur", 30, align_left),
        ("Total RAPOR + TKA", 18, align_right),
        ("Jarak (m)", 15, align_right),
        ("Skor Sertifikat", 15, align_right),
        ("Domisili", 15, align_center),
        ("Skor Akhir", 15, align_right)
    ]
    
    pathways = school_data.get('pathways', {})
    
    for key, applicants in pathways.items():
        name = key.capitalize()
        # Clean name mapping
        name_map = {
            'afirmasi': 'Afirmasi',
            'domisili': 'Domisili',
            'mutasi': 'Mutasi',
            'prestasi': 'Prestasi'
        }
        sheet_title = name_map.get(key, name)
        
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        # Write Title block
        ws.cell(row=1, column=1, value=f"DAFTAR PENDAFTAR - JALUR {sheet_title.upper()}").font = title_font
        ws.cell(row=2, column=1, value=f"{school_name} - SPMB 2026").font = Font(name=font_family, size=11, italic=True)
        ws.cell(row=3, column=1, value=f"Total Terverifikasi: {len(applicants)} murid").font = Font(name=font_family, size=10)
        
        # Header Row is at Row 5
        for col_idx, (col_name, _, align) in enumerate(columns, 1):
            cell = ws.cell(row=5, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border
            
        # Helper function for parsing numeric values safely for sorting
        def get_sort_val(app, sort_key, desc):
            # Clean values before sorting
            if sort_key == 'nilai_rapor' and app.get('sub_jalur') == 'Prestasi Akademik/Non Akademik':
                return -999999 if desc else 999999
            if sort_key == 'skor_sertifikat' and app.get('sub_jalur') == 'Prestasi Rapor':
                return -999999 if desc else 999999

            val = app.get(sort_key)
            if val is None or val == '-':
                return -999999 if desc else 999999
            try:
                return float(val)
            except ValueError:
                return val
                
        # Sort applicants based on pathway
        sorted_applicants = list(applicants)
        if key == 'prestasi':
            # Sort by Nilai Rapor descending, then Skor Sertifikat descending
            sorted_applicants.sort(
                key=lambda x: (
                    get_sort_val(x, 'nilai_rapor', True),
                    get_sort_val(x, 'skor_sertifikat', True)
                ),
                reverse=True
            )
        elif key == 'domisili' or key == 'afirmasi' or key == 'mutasi':
            # Sort by Jarak ascending
            sorted_applicants.sort(
                key=lambda x: get_sort_val(x, 'jarak', False)
            )
            
        # Write data rows starting at Row 6
        for row_idx, app in enumerate(sorted_applicants, 6):
            sub_j = app.get('sub_jalur', '-')
            
            # No
            ws.cell(row=row_idx, column=1, value=row_idx - 5).alignment = align_center
            
            # No. Pendaftaran
            ws.cell(row=row_idx, column=2, value=app.get('no_pendaftaran', '-')).alignment = align_center
            
            # Nama Murid
            ws.cell(row=row_idx, column=3, value=app.get('nama_murid', '-')).alignment = align_left
            
            # Asal Sekolah
            ws.cell(row=row_idx, column=4, value=app.get('asal_satuan_pendidikan', '-')).alignment = align_left
            
            # Sub Jalur
            ws.cell(row=row_idx, column=5, value=sub_j).alignment = align_left
            
            # Nilai Rapor (Total RAPOR + TKA)
            rapor_val = '-'
            if sub_j != 'Prestasi Akademik/Non Akademik':
                raw_rapor = app.get('nilai_rapor', '-')
                if raw_rapor != '-':
                    try:
                        rapor_val = float(raw_rapor)
                    except ValueError:
                        rapor_val = raw_rapor
            ws.cell(row=row_idx, column=6, value=rapor_val).alignment = align_right
            
            # Jarak (m)
            jarak_val = app.get('jarak', '-')
            if jarak_val != '-':
                try:
                    jarak_val = float(jarak_val)
                except ValueError:
                    pass
            ws.cell(row=row_idx, column=7, value=jarak_val).alignment = align_right
            
            # Skor Sertifikat
            sert_val = '-'
            if sub_j != 'Prestasi Rapor':
                raw_sert = app.get('skor_sertifikat', '-')
                if raw_sert != '-':
                    try:
                        sert_val = float(raw_sert)
                    except ValueError:
                        sert_val = raw_sert
            ws.cell(row=row_idx, column=8, value=sert_val).alignment = align_right
            
            # Domisili
            ws.cell(row=row_idx, column=9, value=app.get('status_domisili', '-')).alignment = align_center
            
            # Skor Akhir
            akhir_val = app.get('skor_akhir', '-')
            if akhir_val != '-':
                try:
                    akhir_val = float(akhir_val)
                except ValueError:
                    pass
            ws.cell(row=row_idx, column=10, value=akhir_val).alignment = align_right
            
            # Set styling for data cells
            for col_idx in range(1, 11):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Apply number formatting if cells are float
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'
                    
                # Format integer rank
                if col_idx == 1:
                    cell.font = bold_data_font

        # Adjust column widths
        for col_idx, (_, width, _) in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            
    wb.save(output_file)
    print(f"Excel file created successfully: {output_file}")

def main():
    json_path = Path('pendaftar_data.json')
    if not json_path.exists():
        print("Error: pendaftar_data.json not found.")
        return
        
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    # Clean up the old file name if it exists to avoid confusion
    old_file = Path("SPMB_SMPN1_Bogor_2026_Terurut.xlsx")
    
    schools = data.get('schools', {})
    for key, school_data in schools.items():
        output_file = f"SPMB_{key.upper()}_Bogor_2026_Terurut.xlsx"
        generate_excel_for_school(school_data, output_file)

if __name__ == '__main__':
    main()
